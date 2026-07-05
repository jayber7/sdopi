#!/usr/bin/env python3
"""Extract rubros+items from PLANILLAS and CAO quantities from PLANILLAS
CAO MONTO columns (deriving quantity = MONTO / PU) for CAO N°1-7.

Output: frontend/tests/fixtures/cao-data-extracted.json + diff report."""

import json
import re
from pathlib import Path

from openpyxl import load_workbook

EXCEL = "Docs/ANALISIS DE CAO ROSARIO DEL INGRE V4.xlsm"
OUT = "frontend/tests/fixtures/cao-data-extracted.json"
EXISTING = "frontend/tests/fixtures/cao-data.json"
NUM_CAOS = 7

SRC_DIR = Path(__file__).resolve().parent.parent
excel_path = SRC_DIR / EXCEL
out_path = SRC_DIR / OUT
existing_path = SRC_DIR / EXISTING

wb = load_workbook(str(excel_path), data_only=True)
ws = wb['PLANILLAS']

# PLANILLAS columns (0-indexed)
# Row 12: ITEM | DESCRIPCION | SEGÚN CONTRATO |   |   |   | OT1 |   |   |   | CAO1 |   |   | CAO2 | ...
# Row 13:      |             | PU |   |CANT|MONTO| ...  sub-headers
# Row 14:      |             |UNID|Bs.|    |     | ...
# Data starts row 15 (rubro headers) / row 16 (items)
# Each CAO block: CANTIDAD(col 0) | MONTO(col+1) | AVANCE%(col+2)
CAO_CANT_START = 10  # 0-indexed col where CAO 1 CANTIDAD starts
CAO_SPAN = 3

RUBRO_RE = re.compile(r'^(M\d{2})\s*-\s*(.*)')

rubros = []
items = []
current_rubro = None

for r in range(15, ws.max_row + 1):
    nro = ws.cell(r, 1).value       # col 0
    desc = ws.cell(r, 2).value       # col 1
    unidad = ws.cell(r, 3).value     # col 2
    pu = ws.cell(r, 4).value         # col 3
    cc = ws.cell(r, 5).value         # col 4

    if isinstance(desc, str) and (m := RUBRO_RE.match(desc.strip())):
        current_rubro = {"codigo": m.group(1), "nombre": m.group(2).strip(), "items": []}
        rubros.append(current_rubro)
        continue

    if not isinstance(nro, (int, float)):
        continue
    if not isinstance(pu, (int, float)) or pu <= 0:
        continue

    n = int(nro)
    pu_f = float(pu)
    cc_f = float(cc) if isinstance(cc, (int, float)) else 0.0

    item = {
        "numero": n,
        "descripcion": (desc or "").strip(),
        "unidad": (unidad or "").strip(),
        "precioUnitario": pu_f,
        "cantidadContrato": cc_f,
    }
    items.append(item)
    if current_rubro is not None:
        current_rubro["items"].append(item)

print(f"PLANILLAS: {len(rubros)} rubros, {len(items)} items")

# ── CAO quantities: MONTO / PU ────────────────────────────────────
# Each CAO spans 3 cols: CANTIDAD(offset 0) | MONTO(offset 1) | AVANCE%(offset 2)
caos = {}
for cao_n in range(1, NUM_CAOS + 1):
    monto_col = CAO_CANT_START + (cao_n - 1) * CAO_SPAN + 1  # MONTO column (0-idx)
    cao_list = []
    item_idx = 0
    for r in range(15, ws.max_row + 1):
        nro = ws.cell(r, 1).value
        if not isinstance(nro, (int, float)):
            continue
        monto = ws.cell(r, monto_col + 1).value  # 1-indexed column
        if isinstance(monto, (int, float)) and monto > 0:
            pu_item = items[item_idx]["precioUnitario"]
            cantidad = round(monto / pu_item, 4) if pu_item > 0 else 0
            if cantidad > 0:
                cao_list.append({"itemNumero": items[item_idx]["numero"], "cantidad": cantidad})
        item_idx += 1
    caos[str(cao_n)] = cao_list
    print(f"  CAO {cao_n}: {len(cao_list)} items")

wb.close()

# ── Write output ──────────────────────────────────────────────────
output = {
    "proyecto": {
        "nombre": "MEJORAMIENTO CAMINO ROSARIO DEL INGRE - MACHICOCA - IVOCA (FASE I)",
        "contratoNro": "005/2019",
        "montoContrato": 16903840.54,
        "ordenProceder": "2019-10-21",
        "fechaConclusion": "2022-01-09",
        "direccion": "Rosario del Ingre - Machicoca, Provincia Cordillera",
        "contratista": "CONSORCIO ROSARIO",
        "supervisor": "SUPERVISORA CAMINOS",
        "fiscal": "DIRECCION DE INFRAESTRUCTURA",
        "anticipoPct": 13.7747448,
    },
    "rubros": rubros,
    "caos": caos,
}

out_path.write_text(json.dumps(output, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
print(f"\n✓ Written:  {OUT}")

# ── Diff vs existing ─────────────────────────────────────────────
if not existing_path.exists():
    print(f"⚠ {EXISTING} not found — skipping diff.")
    sys.exit(0)

import sys

existing = json.loads(existing_path.read_text(encoding="utf-8"))
diffs = []

# Structure
if len(output["rubros"]) != len(existing["rubros"]):
    diffs.append(f"rubro count: {len(output['rubros'])} vs {len(existing['rubros'])}")

e_per_rubro = {r["codigo"]: len(r["items"]) for r in output["rubros"]}
x_per_rubro = {r["codigo"]: len(r["items"]) for r in existing["rubros"]}
all_codes = set(list(e_per_rubro.keys()) + list(x_per_rubro.keys()))
for c in sorted(all_codes):
    if e_per_rubro.get(c, 0) != x_per_rubro.get(c, 0):
        diffs.append(f"  {c}: {e_per_rubro.get(c,0)} items vs {x_per_rubro.get(c,0)}")

e_total = sum(len(r["items"]) for r in output["rubros"])
x_total = sum(len(r["items"]) for r in existing["rubros"])
if e_total != x_total:
    diffs.append(f"total items: {e_total} vs {x_total}")

# Item content (only for items that exist in both)
x_item_map = {it["numero"]: it for r in existing["rubros"] for it in r["items"]}
item_diffs = []
for r in output["rubros"]:
    for it in r["items"]:
        xi = x_item_map.get(it["numero"])
        if xi is None:
            item_diffs.append(f"  item {it['numero']}: (new — not in existing)")
            continue
        for k in ("precioUnitario", "cantidadContrato", "descripcion", "unidad"):
            if k == "descripcion":
                if it[k].replace('"', '').replace("'", '') != xi[k].replace('"', '').replace("'", ''):
                    item_diffs.append(f"  item {it['numero']} {k}: '{it[k]}' vs '{xi[k]}'")
            elif isinstance(it[k], float) and abs(it[k] - xi[k]) > 0.02:
                item_diffs.append(f"  item {it['numero']} {k}: {it[k]} vs {xi[k]}")
            elif not isinstance(it[k], float) and it[k] != xi[k]:
                item_diffs.append(f"  item {it['numero']} {k}: {it[k]} vs {xi[k]}")

# CAO quantities
cao_diffs = []
for cao_k in sorted(output["caos"], key=int):
    e_map = {e["itemNumero"]: e["cantidad"] for e in output["caos"][cao_k]}
    x_map = {x["itemNumero"]: x["cantidad"] for x in existing["caos"].get(cao_k, [])}
    for n in sorted(set(list(e_map.keys()) + list(x_map.keys()))):
        ev = e_map.get(n)
        xv = x_map.get(n)
        if ev is None:
            cao_diffs.append(f"  CAO {cao_k} item {n}: missing in extracted, existing={xv}")
        elif xv is None:
            cao_diffs.append(f"  CAO {cao_k} item {n}: extracted={ev}, missing in existing")
        elif abs(ev - xv) > 0.02:
            cao_diffs.append(f"  CAO {cao_k} item {n}: extracted={ev}, existing={xv}")

print(f"\n── DIFF vs {EXISTING} ──")
if not diffs and not item_diffs and not cao_diffs:
    print("✅ Full match (no differences found)")
elif not diffs and not cao_diffs and len(item_diffs) == len(set(it["numero"] for r in output["rubros"] for it in r["items"]) - set(x_item_map.keys())):
    # Only diffs are new items — clean
    print("✅ Match except for new items:")
    for d in item_diffs:
        print(d)
else:
    if diffs:
        print(f"Structure ({len(diffs)}):")
        for d in diffs:
            print(d)
    if item_diffs:
        print(f"Item content ({len(item_diffs)}):")
        for d in item_diffs:
            print(d)
    if cao_diffs:
        print(f"CAO quantities ({len(cao_diffs)}):")
        for d in cao_diffs:
            print(d)
    if not diffs and not item_diffs and not cao_diffs:
        print("✅ Full match")

print("\nDone.")
